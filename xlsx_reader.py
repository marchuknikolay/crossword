"""Read and validate crossword clues from an XLSX workbook."""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

from models import ClueEntry, CrosswordError


def read_clues(path: str | Path, grid_size: int = 15) -> list[ClueEntry]:
    """Open *path*, detect header, parse rows, validate and return clue entries."""
    path = Path(path)
    if not path.exists():
        raise CrosswordError(f"File not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    header_row = _detect_header_row(ws)
    entries: list[ClueEntry] = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[0] is None:
            continue
        try:
            number = int(row[0])
        except (ValueError, TypeError):
            continue
        clue_text = str(row[1]) if row[1] else ""
        raw_answer = str(row[2]) if row[2] else ""
        answer = _normalize_answer(raw_answer)
        if not answer:
            continue
        entries.append(ClueEntry(number=number, clue_text=clue_text, answer=answer))

    wb.close()
    return _validate_and_filter(entries, grid_size)


def _detect_header_row(sheet) -> int:
    """Return the 1-based row index of the first row where column A is an int.

    The row before that is assumed to be the header.  Falls back to row 1.
    """
    for row in sheet.iter_rows(min_row=1, max_row=20, max_col=1, values_only=False):
        cell = row[0]
        try:
            int(cell.value)
            # This row is data; header is the row before
            return max(1, cell.row - 1)
        except (ValueError, TypeError):
            continue
    return 1


def _normalize_answer(raw: str) -> str:
    """Uppercase, strip everything except A-Z."""
    return "".join(c for c in raw.upper() if c.isalpha())


def _validate_and_filter(
    entries: list[ClueEntry], grid_size: int
) -> list[ClueEntry]:
    """Keep answers of length 3..grid_size, deduplicate, error if none remain."""
    seen_answers: set[str] = set()
    result: list[ClueEntry] = []

    for entry in entries:
        if len(entry.answer) < 3:
            print(
                f"Warning: skipping '{entry.answer}' (too short, <3 letters)",
                file=sys.stderr,
            )
            continue
        if len(entry.answer) > grid_size:
            print(
                f"Warning: skipping '{entry.answer}' (too long for {grid_size}x{grid_size} grid)",
                file=sys.stderr,
            )
            continue
        if entry.answer in seen_answers:
            print(
                f"Warning: duplicate answer '{entry.answer}', skipping",
                file=sys.stderr,
            )
            continue
        seen_answers.add(entry.answer)
        result.append(entry)

    if not result:
        raise CrosswordError("No valid clue entries after filtering")

    return result
