"""Tests for xlsx_writer.py."""

import os
import tempfile

import openpyxl
import pytest

from models import ClueEntry, Direction, NumberedClue
from xlsx_writer import write_clues_xlsx


def _sample_clues():
    across = [
        NumberedClue(1, "Feline pet", "CAT", Direction.ACROSS),
        NumberedClue(5, "Man's best friend", "DOG", Direction.ACROSS),
    ]
    down = [
        NumberedClue(1, "Automobile", "CAR", Direction.DOWN),
        NumberedClue(3, "Large body of water", "OCEAN", Direction.DOWN),
    ]
    return across, down


class TestWriteCluesXlsx:
    def test_creates_valid_xlsx(self):
        across, down = _sample_clues()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            write_clues_xlsx(across, down, path)
            assert os.path.exists(path)
            wb = openpyxl.load_workbook(path)
            assert "Clues" in wb.sheetnames
        finally:
            os.unlink(path)

    def test_across_section(self):
        across, down = _sample_clues()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            write_clues_xlsx(across, down, path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            assert ws.cell(row=1, column=1).value == "ACROSS"
            assert ws.cell(row=2, column=1).value == "1. Feline pet"
            assert ws.cell(row=2, column=2).value == "CAT"
            assert ws.cell(row=3, column=1).value == "5. Man's best friend"
            assert ws.cell(row=3, column=2).value == "DOG"
        finally:
            os.unlink(path)

    def test_down_section(self):
        across, down = _sample_clues()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            write_clues_xlsx(across, down, path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            # Row 4 is blank separator, row 5 is DOWN header
            assert ws.cell(row=5, column=1).value == "DOWN"
            assert ws.cell(row=6, column=1).value == "1. Automobile"
            assert ws.cell(row=6, column=2).value == "CAR"
            assert ws.cell(row=7, column=1).value == "3. Large body of water"
            assert ws.cell(row=7, column=2).value == "OCEAN"
        finally:
            os.unlink(path)

    def test_blank_separator_row(self):
        across, down = _sample_clues()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            write_clues_xlsx(across, down, path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            # Row 4 should be blank (separator between ACROSS and DOWN)
            assert ws.cell(row=4, column=1).value is None
        finally:
            os.unlink(path)

    def test_bold_headers(self):
        across, down = _sample_clues()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            write_clues_xlsx(across, down, path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            assert ws.cell(row=1, column=1).font.bold is True
            assert ws.cell(row=5, column=1).font.bold is True
        finally:
            os.unlink(path)

    def test_empty_clues(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            write_clues_xlsx([], [], path)
            assert os.path.exists(path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            assert ws.cell(row=1, column=1).value == "ACROSS"
            assert ws.cell(row=3, column=1).value == "DOWN"
        finally:
            os.unlink(path)

    def test_unplaced_sheet_created(self):
        across, down = _sample_clues()
        unplaced = [
            ClueEntry(10, "Not used clue", "UNUSED"),
            ClueEntry(11, "Another skipped", "SKIPPED"),
        ]
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            write_clues_xlsx(across, down, path, unplaced=unplaced)
            wb = openpyxl.load_workbook(path)
            assert "Not placed" in wb.sheetnames
            ws2 = wb["Not placed"]
            assert ws2.cell(row=1, column=1).value == "Clue"
            assert ws2.cell(row=1, column=2).value == "Answer"
            assert ws2.cell(row=2, column=1).value == "Not used clue"
            assert ws2.cell(row=2, column=2).value == "UNUSED"
            assert ws2.cell(row=3, column=1).value == "Another skipped"
            assert ws2.cell(row=3, column=2).value == "SKIPPED"
        finally:
            os.unlink(path)

    def test_no_unplaced_sheet_when_empty(self):
        across, down = _sample_clues()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            write_clues_xlsx(across, down, path, unplaced=[])
            wb = openpyxl.load_workbook(path)
            assert "Not placed" not in wb.sheetnames
        finally:
            os.unlink(path)

    def test_no_unplaced_sheet_when_none(self):
        across, down = _sample_clues()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            write_clues_xlsx(across, down, path)
            wb = openpyxl.load_workbook(path)
            assert "Not placed" not in wb.sheetnames
        finally:
            os.unlink(path)
