"""Write placed crossword clues to an XLSX file."""

from __future__ import annotations

import openpyxl
from openpyxl.styles import Font

from models import ClueEntry, NumberedClue


def write_clues_xlsx(
    across: list[NumberedClue],
    down: list[NumberedClue],
    output_path: str,
    unplaced: list[ClueEntry] | None = None,
) -> None:
    """Write across and down clues to an Excel workbook.

    Numbering is embedded in the clue cell: '1. Clue text'.
    Answers are in column B.
    If *unplaced* is provided, a second sheet lists words that didn't fit.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clues"

    header_font = Font(bold=True, size=12)
    row = 1

    # ACROSS section
    ws.cell(row=row, column=1, value="ACROSS").font = header_font
    row += 1
    for clue in across:
        ws.cell(row=row, column=1, value=f"{clue.number}. {clue.clue_text}")
        ws.cell(row=row, column=2, value=clue.answer)
        row += 1

    # Blank separator
    row += 1

    # DOWN section
    ws.cell(row=row, column=1, value="DOWN").font = header_font
    row += 1
    for clue in down:
        ws.cell(row=row, column=1, value=f"{clue.number}. {clue.clue_text}")
        ws.cell(row=row, column=2, value=clue.answer)
        row += 1

    # Set column widths
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 15

    # Unplaced words sheet
    if unplaced:
        ws2 = wb.create_sheet(title="Not placed")
        ws2.cell(row=1, column=1, value="Clue").font = header_font
        ws2.cell(row=1, column=2, value="Answer").font = header_font
        for i, clue in enumerate(unplaced, start=2):
            ws2.cell(row=i, column=1, value=clue.clue_text)
            ws2.cell(row=i, column=2, value=clue.answer)
        ws2.column_dimensions["A"].width = 60
        ws2.column_dimensions["B"].width = 15

    wb.save(output_path)
